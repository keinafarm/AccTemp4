# -*- coding: utf-8 -*-
import openpyxl
import datetime
import re

"""

    Excel File操作部

"""


class NoneCell:
    """
    valueに何をセットされても、常にNoneな値を返すCell
    """

    @property
    def value(self):
        return None

    @value.setter
    def value(self, arg):
        pass


class ATCell:
    """
    valueが抱えているcellのvalueに見えるクラス
    """

    def __init__(self, cell):
        self.cell = cell

    @property
    def value(self):
        return self.cell.value

    @value.setter
    def value(self, value):
        self.cell.value = value


class ATRow:
    """
    Excelのセルデータと、本システム内部の管理情報の構造を橋渡しする
    """

    def __init__(self, row_data, column_position):
        self.row = row_data
        self.column_position = column_position
        self.cell_data = [None, None, None, None, None, None]
        for i in range(0, 6):
            if column_position[i] is None:
                cell = NoneCell()  # 該当カラムが存在しない時は、常にNoneを返す
            else:
                cell = row_data[column_position[i] - 1]
            cell_object = ATCell(cell)
            self.cell_data[i] = cell_object

    @property
    def start_date(self):
        return self.cell_data[0].value

    @start_date.setter
    def start_date(self, value):
        self.cell_data[0].value = value

    @property
    def end_date(self):
        return self.cell_data[1].value

    @end_date.setter
    def end_date(self, value):
        self.cell_data[1].value = value

    @property
    def target_temperature(self):
        return self.cell_data[2].value

    @target_temperature.setter
    def target_temperature(self, value):
        self.cell_data[2].value = value

    @property
    def current_temperature(self):
        return self.cell_data[3].value

    @current_temperature.setter
    def current_temperature(self, value):
        self.cell_data[3].value = value

    @property
    def rate(self):
        return self.cell_data[4].value

    @rate.setter
    def rate(self, value):
        self.cell_data[4].value = value

    @property
    def estimate_date(self):
        return self.cell_data[5].value

    @estimate_date.setter
    def estimate_date(self, value):
        self.cell_data[5].value = value

    def rowNo(self):
        return self.row[0].row

    def __str__(self):
        result = "開始日:" + (
            "空白" if self.start_date is None else self.start_date.strftime('%m/%d'))
        result += "  終了日:" + (
            "空白" if self.end_date is None else self.end_date.strftime('%m/%d'))
        result += "  目標積算温度:" + (
            "空白" if self.target_temperature is None else str(self.target_temperature))
        result += "  現状積算温度:" + (
            "空白" if self.current_temperature is None else str(self.current_temperature))
        result += "  目標到達度:" + (
            "空白" if self.rate is None else str(self.rate))
        result += "  予測終了日:" + (
            "空白" if self.estimate_date is None else self.estimate_date.strftime(
                '%m/%d'))
        return result


class ATSheet:
    def __init__(self, work_sheet):
        """
        積算温度シートクラス

        ２行目に、開始日", "終了日", "目標積算温度", "現状積算温度", "目標到達度", "予測終了日"の各タイトルが出てくるカラム位置を記憶する

        :param work_sheet: ワークシート
        """
        self.title_name_start_date = None
        self.title_name_end_date = None
        self.title_name_target_temperature = None
        self.title_name_current_temperature = None
        self.title_name_rate = None
        self.title_name_estimate_date = None
        self.title_name_line_no = 0
        self.column_position = None
        self.rows = None
        self.work_sheet = work_sheet

    def get_column_value_string(self, row_no):
        """
        指定した行番号のデータを得る
        :param row_no:行番号
        :return:セルの値リスト
        """
        data_list = [str(cell[0].value) for cell in self.work_sheet.iter_cols(min_row=row_no, max_row=row_no)]
        return data_list

    def set_column_position_string(self, line_no, start_date, end_date, target_temperature, current_temperature, rate,
                                   estimate_date):
        """
        開始日、終了日、目標積算温度、現在の積算温度、達成率、予測終了日の各データが格納されているタイトル名を指定する
        :param line_no: タイトル行の番号
        :param start_date: 開始日のタイトル名
        :param end_date: 終了日のタイトル名
        :param target_temperature: 目標積算温度のタイトル名
        :param current_temperature: 現在の積算温度のタイトル名
        :param rate: 達成率のタイトル名
        :param estimate_date: 予測終了日のタイトル名
        :return:
        """

        # 指定されたカラム名と同じ文字列を見つけたら、その位置を記憶する
        self.column_position = [None, None, None, None, None, None]  # 記憶する場所
        # タイトル名として比較する文字列のリスト
        compare_list = [start_date, end_date, target_temperature, current_temperature, rate, estimate_date]

        for cell in self.work_sheet.iter_cols(min_row=line_no, max_row=line_no):
            value = cell[0].value
            index = compare_list.index(value) if value in compare_list else None
            if index is not None:
                self.column_position[index] = cell[0].column

        self.rows = []
        for row_data in self.work_sheet.iter_rows(min_row=3):
            row_obj = ATRow(row_data, self.column_position)
            self.rows.append(row_obj)

    @property
    def data_list(self):
        return self.rows


class ATWorkBook:
    def __init__(self, file_name):
        """
        積算温度ワークブッククラス
        :param file_name: エクセルファイル名
        """
        self.file_name = file_name
        self.work_book = openpyxl.load_workbook(file_name)
        """
    開始日(start_date)
    終了日(end_date)
    目標積算温度(target_temperature)
    現状積算温度(current_temperature)
    目標到達度(rate)
    予測終了日(estimate_date)
        """
        self.target_sheets = []
        for ws in self.work_book:  # work sheetの名前リストでループ
            if ws.title.find("積算温度") == 0:
                ws_obj = ATSheet(ws)
                self.target_sheets.append(ws_obj)

        print(self.target_sheets)

    def search_sheets(self, pattern):
        """
        指定した正規表現に一致するシート名を対象にする
        :param pattern: 検索する正規表現文字列
        :return:
        """

        self.target_sheets = []
        p = re.compile(pattern)
        for ws in self.work_book:  # work sheetの名前リストでループ
            if p.match(ws.title) is not None:
                ws_obj = ATSheet(ws)
                self.target_sheets.append(ws_obj)

    def target_sheet_all(self):
        """
        すべてのシートを対象にする
        :return:
        """
        self.target_sheets = [ATSheet(ws) for ws in self.work_book]

    def set_target_sheets_by_name_list(self, sheet_name_list):
        """
        指定されたシート名リストに従って、ワークシートオブジェクトリストを取得する
        :param sheet_name_list: シート名リスト
        :return:ワークシートオブジェクトリスト
        """
        self.target_sheets = [ATSheet(self.work_book.get_sheet_by_name(name)) for name in sheet_name_list]

    def flash(self):
        """
        データを上書きする（今回の仕様では使ってない）
        :return:
        """
        self.work_book.save(self.file_name)

    def save(self, file_name):
        """
        指定したファイルに出力する
        :param file_name: 出力するファイル名
        :return:
        """
        self.work_book.save(file_name)

    def get_row_data_string(self, line_no):
        """
        指定した行の各セルのデータを文字列として取得する
        :param line_no:
        :return:
        """
        row_data = []
        for ws in self.target_sheets:
            data_in_sheet = ws.get_column_value_string(line_no)
            row_data.append(data_in_sheet)
        return row_data

    def set_target_work_sheet(self, work_sheet_name_list):
        """
        対象とするシートをシート名のリストで指定する
        :param work_sheet_name_list: シート名リスト
        :return:
        """
        work_sheet_name_list = list(work_sheet_name_list)
        self.target_sheets = []
        for ws in self.work_book:  # work sheetの名前リストでループ
            if ws.title in work_sheet_name_list:
                ws_obj = ATSheet(ws)
                self.target_sheets.append(ws_obj)

    def set_column_position_string(self, line_no, start_date, end_date, target_temperature, current_temperature, rate,
                                   estimate_date):
        """
        今回、操作するデータの対象となるカラムを、指定した行をタイトル名として、選択する
        :param line_no: タイトル名の行番号
        :param start_date: 開始日のタイトル名
        :param end_date: 終了日のタイトル名
        :param target_temperature: 目標積算温度のタイトル名
        :param current_temperature: 現状積算温度のタイトル名
        :param rate: 目標達成率のタイトル名
        :param estimate_date: 予測終了日 のタイトル名
        :return:
        """
        for ws in self.target_sheets:
            ws.set_column_position_string(line_no, start_date, end_date, target_temperature, current_temperature, rate,
                                          estimate_date)

    @property
    def sheets_list(self):
        return self.target_sheets

    @property
    def sheets_name_list(self):
        name_list = [ws.work_sheet.title for ws in self.target_sheets]
        return name_list


if __name__ == "__main__":
    try:
        obj = ATWorkBook("test1.xlsx")
    except FileNotFoundError as e:
        print(e)
        print(type(e))
    else:
        obj.set_column_position_string(2, "開始日", "終了日", "目標積算温度", "現状積算温度", "目標到達度", "予測終了日")
        for sheet in obj.sheets_list:
            for data in sheet.data_list:
                print(data)

        row = obj.sheets_list[0].data_list[2]
        row.target_temperature = 256
        row.current_temperature = 512
        row.rate = 20.3
        row.estimate_date = datetime.datetime(2020, 12, 24)

        row = obj.sheets_list[0].data_list[0]
        row.start_date = datetime.datetime(2020, 2, 24)
        row.end_date = datetime.datetime(2020, 3, 24)
        obj.flash()
        print("Done")
